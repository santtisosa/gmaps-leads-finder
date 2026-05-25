"""
enviar_whatsapp.py
Envía mensajes de WhatsApp personalizados a cada lead en leads.xlsx.

Requisitos:
  - Chrome/Chromium instalado
  - WhatsApp Web abierto y logueado en el navegador por defecto
  - leads.xlsx generado por buscar_leads.py

Uso:
  venv/bin/python enviar_whatsapp.py
  venv/bin/python enviar_whatsapp.py --csv leads.xlsx --delay 30 --limite 20
"""

import time
import argparse
import re
import os
import yaml
import openpyxl
import pywhatkit as kit
from dotenv import load_dotenv

load_dotenv()

# ── CONFIG ──────────────────────────────────────────────────────────────────
_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.yaml")

def _load_config(path=_CONFIG_PATH):
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)

CFG = _load_config()

CSV_FILE         = CFG["archivos"]["output"]
DELAY_SEG        = CFG["whatsapp"]["delay_seg"]
WAIT_TIME        = CFG["whatsapp"]["wait_time"]
CLOSE_TIME       = CFG["whatsapp"]["close_time"]
MENSAJE_TEMPLATE = CFG["mensaje"]
LIMITE           = None
# ────────────────────────────────────────────────────────────────────────────


def limpiar_numero(telefono):
    """
    Convierte un teléfono a formato internacional sin + ni espacios.
    Asume Uruguay (+598) si no tiene código de país.
    """
    digits = re.sub(r"[^\d+]", "", telefono)

    if digits.startswith("+"):
        return digits
    elif digits.startswith("598"):
        return "+" + digits
    elif len(digits) >= 8:
        return "+598" + digits.lstrip("0")
    return None


def parse_args():
    parser = argparse.ArgumentParser(description="Envía mensajes WA a leads del Excel")
    parser.add_argument("--csv",    default=CSV_FILE,  help=f"Archivo Excel (default: {CSV_FILE})")
    parser.add_argument("--delay",  default=DELAY_SEG, type=int, help=f"Segundos entre mensajes (default: {DELAY_SEG})")
    parser.add_argument("--limite", default=LIMITE,    type=int, help="Máximo de mensajes a enviar (default: todos)")
    parser.add_argument("--dry-run", action="store_true", help="Muestra qué haría pero no envía nada")
    parser.add_argument("--config", default=_CONFIG_PATH, help=f"Ruta al config.yaml (default: {_CONFIG_PATH})")
    return parser.parse_args()


def cargar_leads(xlsx_path):
    """Lee el Excel y devuelve leads con teléfono no contactados."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if data.get("contactado", "") and str(data["contactado"]).strip():
            continue
        if not data.get("teléfono", "") and not data.get("telefono", ""):
            continue
        data["telefono"] = data.get("teléfono") or data.get("telefono") or ""
        data["nombre"]   = data.get("nombre", "")
        leads.append(data)
    return leads


def marcar_contactado(xlsx_path, nombre):
    """Marca un lead como contactado en el Excel."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    contactado_col = None
    nombre_col     = None
    for i, h in enumerate(headers, 1):
        if h in ("contactado",):   contactado_col = i
        if h == "nombre":          nombre_col = i
    if not contactado_col or not nombre_col:
        return
    for row in ws.iter_rows(min_row=2):
        if str(row[nombre_col - 1].value).strip() == nombre:
            row[contactado_col - 1].value = "✓"
            break
    wb.save(xlsx_path)


def main():
    args = parse_args()

    # Reload config if a custom path was given
    if args.config != _CONFIG_PATH:
        global CFG, CSV_FILE, DELAY_SEG, WAIT_TIME, CLOSE_TIME, MENSAJE_TEMPLATE
        CFG              = _load_config(args.config)
        CSV_FILE         = CFG["archivos"]["output"]
        DELAY_SEG        = CFG["whatsapp"]["delay_seg"]
        WAIT_TIME        = CFG["whatsapp"]["wait_time"]
        CLOSE_TIME       = CFG["whatsapp"]["close_time"]
        MENSAJE_TEMPLATE = CFG["mensaje"]

    if not os.path.exists(args.csv):
        print(f"ERROR: No se encontró {args.csv}. Corré primero buscar_leads.py")
        return

    leads = cargar_leads(args.csv)
    print(f"Leads sin contactar: {len(leads)}")

    if not leads:
        print("No hay leads pendientes.")
        return

    if args.limite:
        leads = leads[:args.limite]
        print(f"Límite: {args.limite} mensajes")

    if args.dry_run:
        print("\n[DRY RUN — no se envía nada]\n")
        for lead in leads:
            numero = limpiar_numero(lead["telefono"])
            mensaje = MENSAJE_TEMPLATE.format(nombre=lead["nombre"])
            print(f"  → {lead['nombre']}")
            print(f"     Tel: {numero}")
            print(f"     Msg: {mensaje[:80]}...")
            print()
        return

    print("\n⚠  Asegurate de tener WhatsApp Web abierto y logueado en Chrome.")
    input("   Presioná ENTER cuando estés listo...\n")

    enviados = 0
    errores  = 0

    for i, lead in enumerate(leads, 1):
        nombre  = lead["nombre"]
        numero  = limpiar_numero(lead["telefono"])
        mensaje = MENSAJE_TEMPLATE.format(nombre=nombre)

        if not numero:
            print(f"[{i}] ⚠  {nombre} — número inválido: {lead['telefono']}")
            errores += 1
            continue

        print(f"[{i}/{len(leads)}] Enviando a {nombre} ({numero})...")

        try:
            kit.sendwhatmsg_instantly(
                phone_no=numero,
                message=mensaje,
                wait_time=WAIT_TIME,
                tab_close=True,
                close_time=CLOSE_TIME,
            )
            marcar_contactado(args.csv, nombre)
            enviados += 1
            print(f"  ✓ Enviado")

        except Exception as e:
            print(f"  ✗ Error: {e}")
            errores += 1

        if i < len(leads):
            print(f"  Esperando {args.delay}s...")
            time.sleep(args.delay)

    print(f"\n{'='*50}")
    print(f"Enviados:  {enviados}")
    print(f"Errores:   {errores}")
    print(f"Excel actualizado: columna 'contactado' marcada con ✓")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
