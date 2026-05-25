"""
buscar_leads.py
Busca negocios en Montevideo sin sitio web usando Google Maps Places API.
Exporta resultados a leads.csv
"""

import os
import csv
import time
import json
import difflib
import unicodedata
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import googlemaps
from datetime import datetime
from dotenv import load_dotenv
from tqdm import tqdm

load_dotenv()  # carga .env automáticamente si existe

# ── CONFIG ─────────────────────────────────────────────────────────────────
API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "TU_API_KEY_AQUI")

CIUDAD = "Montevideo, Uruguay"

CATEGORIAS = [
    "peluquería",
    "barbería",
    "taller mecánico",
    "restaurante",
    "panadería",
    "ferretería",
    "veterinaria",
    "gimnasio",
    "lavandería",
    "cerrajería",
    "dentista",
    "kiosco",
    "carnicería",
    "frutería",
    "farmacia",
]

OUTPUT_FILE = "leads.xlsx"
DELAY_BETWEEN_REQUESTS = 0.2   # segundos entre llamadas API
MAX_PAGES_PER_CATEGORIA = 3    # cada página = 20 resultados → max 60 por categoría
MIN_RESENAS = 5                # mínimo de reseñas para considerar el negocio (filtra negocios fantasma)
COSTS_FILE = "costos.json"

# Precios Google Maps Places API (USD)
PRECIO_TEXT_SEARCH   = 0.032   # por llamada a places()
PRECIO_PLACE_DETAIL  = 0.017   # por llamada a place()
CREDITO_MENSUAL      = 200.00  # crédito gratis de Google por mes

FUZZY_UMBRAL = 0.88  # similitud mínima para considerar duplicado (0-1)
# ───────────────────────────────────────────────────────────────────────────


# ── FUZZY DEDUP ─────────────────────────────────────────────────────────────
def _normalizar(texto: str) -> str:
    """Lowercase + strip accents for fuzzy comparison."""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


def es_duplicado(nombre: str, nombres_norm: set, umbral: float = FUZZY_UMBRAL) -> bool:
    """
    Returns True if `nombre` is already represented in `nombres_norm`.
    First does an exact check (O(1)), then fuzzy via difflib (O(n)).
    Catches accent variations: "Peluquería Juan" ≈ "Peluqueria Juan".
    """
    norm = _normalizar(nombre)
    if norm in nombres_norm:
        return True
    matches = difflib.get_close_matches(norm, nombres_norm, n=1, cutoff=umbral)
    return bool(matches)
# ─────────────────────────────────────────────────────────────────────────────


# ── COST TRACKER ────────────────────────────────────────────────────────────
class CostTracker:
    def __init__(self, path=COSTS_FILE):
        self.path = path
        self.mes_actual = datetime.now().strftime("%Y-%m")
        self.data = self._cargar()
        self.calls_text_search  = 0
        self.calls_place_detail = 0

    def _cargar(self):
        if os.path.exists(self.path):
            with open(self.path, encoding="utf-8") as f:
                data = json.load(f)
            # Resetear si cambió el mes
            if data.get("mes") != self.mes_actual:
                print(f"  Nuevo mes ({self.mes_actual}). Reseteando contador mensual.")
                return self._nuevo_mes()
            return data
        return self._nuevo_mes()

    def _nuevo_mes(self):
        return {
            "mes": self.mes_actual,
            "total_text_search":  0,
            "total_place_detail": 0,
            "costo_total_usd":    0.0,
            "corridas": [],
        }

    def registrar_text_search(self, n=1):
        self.calls_text_search += n

    def registrar_place_detail(self, n=1):
        self.calls_place_detail += n

    def costo_esta_corrida(self):
        return (
            self.calls_text_search  * PRECIO_TEXT_SEARCH +
            self.calls_place_detail * PRECIO_PLACE_DETAIL
        )

    def guardar(self, leads_encontrados):
        costo_corrida = self.costo_esta_corrida()
        self.data["total_text_search"]  += self.calls_text_search
        self.data["total_place_detail"] += self.calls_place_detail
        self.data["costo_total_usd"]    = round(
            self.data["costo_total_usd"] + costo_corrida, 4
        )
        self.data["corridas"].append({
            "fecha":         datetime.now().strftime("%Y-%m-%d %H:%M"),
            "text_searches": self.calls_text_search,
            "place_details": self.calls_place_detail,
            "costo_usd":     round(costo_corrida, 4),
            "leads":         leads_encontrados,
        })
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)

    def imprimir_resumen(self):
        costo_corrida  = self.costo_esta_corrida()
        costo_mensual  = self.data["costo_total_usd"] + costo_corrida
        restante       = max(0, CREDITO_MENSUAL - costo_mensual)
        pct_usado      = (costo_mensual / CREDITO_MENSUAL) * 100

        print(f"\n{'─'*50}")
        print(f"  COSTOS API — {self.mes_actual}")
        print(f"{'─'*50}")
        print(f"  Esta corrida:")
        print(f"    Text searches : {self.calls_text_search:>4}  × ${PRECIO_TEXT_SEARCH} = ${self.calls_text_search * PRECIO_TEXT_SEARCH:.3f}")
        print(f"    Place details : {self.calls_place_detail:>4}  × ${PRECIO_PLACE_DETAIL} = ${self.calls_place_detail * PRECIO_PLACE_DETAIL:.3f}")
        print(f"    Subtotal      : ${costo_corrida:.4f}")
        print(f"  Acumulado mes  : ${costo_mensual:.4f} / ${CREDITO_MENSUAL:.2f} ({pct_usado:.1f}%)")
        print(f"  Crédito restante: ${restante:.4f}")
        print(f"  Historial guardado en: {self.path}")
        print(f"{'─'*50}")
# ────────────────────────────────────────────────────────────────────────────


def buscar_categoria(gmaps, categoria, tracker, ciudad=CIUDAD, max_paginas=MAX_PAGES_PER_CATEGORIA):
    """Devuelve lista de place_id para una categoría en la ciudad."""
    place_ids = []
    query = f"{categoria} {ciudad}"
    print(f"  Buscando: {query}")

    try:
        resp = gmaps.places(query=query)
        tracker.registrar_text_search()
    except Exception as e:
        print(f"  ERROR en búsqueda: {e}")
        return place_ids

    paginas = 0
    while resp and paginas < max_paginas:
        for place in resp.get("results", []):
            place_ids.append(place["place_id"])

        paginas += 1
        next_token = resp.get("next_page_token")
        if not next_token:
            break

        time.sleep(2)  # Google requiere ~2s antes de usar next_page_token
        try:
            resp = gmaps.places(query=query, page_token=next_token)
            tracker.registrar_text_search()
        except Exception as e:
            print(f"  ERROR paginación: {e}")
            break

    return place_ids


def obtener_detalle(gmaps, place_id, tracker, reintentos=3):
    """Devuelve detalle del negocio. Reintenta con backoff exponencial."""
    for intento in range(reintentos):
        try:
            result = gmaps.place(
                place_id,
                fields=[
                    "name",
                    "website",
                    "formatted_phone_number",
                    "formatted_address",
                    "rating",
                    "user_ratings_total",
                ],
            )
            tracker.registrar_place_detail()
            return result.get("result", {})
        except Exception as e:
            if intento < reintentos - 1:
                espera = 2 ** intento
                time.sleep(espera)
            else:
                tqdm.write(f"  ERROR detalle {place_id}: {e}")
    return None


def exportar_excel(leads, path):
    """Exporta leads a .xlsx con formato prolijo. Append si ya existe."""
    HEADERS = ["Nombre", "Teléfono", "Dirección", "Categoría",
               "Rating", "Reseñas", "Contactado", "Notas"]
    KEYS    = ["nombre", "telefono", "direccion", "categoria",
               "rating", "reseñas", "contactado", "notas"]

    # Cargar workbook existente o crear uno nuevo
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        # ── Encabezados ──────────────────────────────────────────────────
        header_fill   = PatternFill("solid", fgColor="1400FF")
        header_font   = Font(bold=True, color="FFFFFF", size=11)
        header_align  = Alignment(horizontal="center", vertical="center")
        thin_border   = Border(
            bottom=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin",  color="DDDDDD"),
        )

        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin_border

        ws.row_dimensions[1].height = 28

    # ── Filas de datos ───────────────────────────────────────────────────
    row_fill_alt  = PatternFill("solid", fgColor="F5F5F5")
    data_font     = Font(size=10)
    data_align    = Alignment(vertical="center", wrap_text=False)

    start_row = ws.max_row + 1 if ws.max_row > 1 else 2

    for i, lead in enumerate(leads):
        r = start_row + i
        fill = row_fill_alt if i % 2 == 0 else PatternFill()
        for col, key in enumerate(KEYS, 1):
            cell = ws.cell(row=r, column=col, value=lead.get(key, ""))
            cell.font      = data_font
            cell.alignment = data_align
            cell.fill      = fill
        ws.row_dimensions[r].height = 20

    # ── Anchos de columna ────────────────────────────────────────────────
    col_widths = [35, 18, 50, 18, 8, 10, 14, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(path)


def parse_args():
    parser = argparse.ArgumentParser(description="Busca negocios sin web en Google Maps")
    parser.add_argument("--ciudad",   default=CIUDAD,              help="Ciudad a buscar (default: Montevideo, Uruguay)")
    parser.add_argument("--output",   default=OUTPUT_FILE,          help="Archivo CSV de salida (default: leads.csv)")
    parser.add_argument("--paginas",  default=MAX_PAGES_PER_CATEGORIA, type=int, help="Páginas por categoría (default: 3)")
    parser.add_argument("--categorias", nargs="+", default=CATEGORIAS, help="Categorías a buscar")
    parser.add_argument("--min-resenas",   default=MIN_RESENAS,   type=int,   help="Mínimo de reseñas para incluir un negocio (default: 5)")
    parser.add_argument("--fuzzy-umbral",  default=FUZZY_UMBRAL,  type=float, help=f"Similitud mínima para detectar duplicados (default: {FUZZY_UMBRAL})")
    parser.add_argument("--no-fuzzy",      action="store_true",               help="Desactivar fuzzy matching, usar solo comparación exacta")
    return parser.parse_args()


def main():
    args = parse_args()

    if API_KEY == "TU_API_KEY_AQUI":
        print("ERROR: Configurá tu API key.")
        print("  export GOOGLE_MAPS_API_KEY='tu_key'")
        print("  o creá un archivo .env con GOOGLE_MAPS_API_KEY=tu_key")
        return

    gmaps   = googlemaps.Client(key=API_KEY)
    tracker = CostTracker()
    leads   = []
    vistos  = set()

    # Cargar nombres ya guardados para no duplicar entre corridas
    umbral = args.fuzzy_umbral if not args.no_fuzzy else 1.01  # >1 disables fuzzy
    nombres_norm = set()
    if os.path.exists(args.output):
        try:
            wb_exist = openpyxl.load_workbook(args.output)
            ws_exist = wb_exist.active
            headers  = [c.value for c in ws_exist[1]]
            if "nombre" in headers:
                col_idx = headers.index("nombre") + 1
                for row in ws_exist.iter_rows(min_row=2, values_only=True):
                    val = row[col_idx - 1]
                    if val:
                        nombres_norm.add(_normalizar(str(val)))
            modo = "exacto" if args.no_fuzzy else f"fuzzy umbral={umbral}"
            print(f"Leads existentes en {args.output}: {len(nombres_norm)} (dedup {modo})")
        except Exception:
            pass

    print(f"Ciudad: {args.ciudad}")
    print(f"Categorías: {len(args.categorias)}")
    print(f"Output: {args.output}\n")

    for categoria in args.categorias:
        print(f"\n[{categoria.upper()}]")
        place_ids = buscar_categoria(gmaps, categoria, tracker, ciudad=args.ciudad, max_paginas=args.paginas)
        print(f"  Encontrados: {len(place_ids)} negocios")

        sin_web = 0
        nuevos = [p for p in place_ids if p not in vistos]
        for pid in tqdm(nuevos, desc=f"  {categoria}", unit="neg", leave=False):
            vistos.add(pid)

            time.sleep(DELAY_BETWEEN_REQUESTS)
            detalle = obtener_detalle(gmaps, pid, tracker)
            if not detalle:
                continue

            # Sin website = lead potencial
            nombre = detalle.get("name", "")
            resenas = detalle.get("user_ratings_total") or 0
            if not detalle.get("website") and resenas >= args.min_resenas and not es_duplicado(nombre, nombres_norm, umbral):
                sin_web += 1
                nombres_norm.add(_normalizar(nombre))
                leads.append({
                    "nombre":     nombre,
                    "telefono":   detalle.get("formatted_phone_number", ""),
                    "direccion":  detalle.get("formatted_address", ""),
                    "categoria":  categoria,
                    "rating":     detalle.get("rating", ""),
                    "reseñas":    detalle.get("user_ratings_total", ""),
                    "website":    "",
                    "contactado": "",
                    "notas":      "",
                })

        print(f"  Sin web: {sin_web}")

    # ── EXPORTAR EXCEL ──────────────────────────────────────────────────────
    if not leads:
        print("\nNo se encontraron leads.")
        return

    exportar_excel(leads, args.output)

    print(f"\n{'='*50}")
    print(f"TOTAL LEADS: {len(leads)}")
    print(f"Exportado a: {args.output}")
    print(f"{'='*50}")

    tracker.imprimir_resumen()
    tracker.guardar(len(leads))

    if leads:
        exportar_mensajes(leads)


def generar_mensaje(lead):
    """Genera un mensaje de WhatsApp personalizado para un lead."""
    nombre   = lead["nombre"]
    categoria = lead["categoria"]
    telefono  = lead["telefono"]

    mensaje = (
        f"Hola, vi que {nombre} todavía no tiene sitio web. "
        f"Te puedo armar uno en menos de 14 días, con dominio incluido y sin letra chica. "
        f"Antes de cualquier pago te mando una maqueta gratis para que veas cómo quedaría. "
        f"¿Te interesa? — Santiago | santiagososa.dev"
    )

    # Link de WhatsApp listo para abrir (si hay teléfono)
    numero_limpio = "".join(c for c in telefono if c.isdigit())
    if numero_limpio:
        from urllib.parse import quote
        link = f"https://wa.me/{numero_limpio}?text={quote(mensaje)}"
    else:
        link = ""

    return mensaje, link


def exportar_mensajes(leads, output="mensajes.txt"):
    """Exporta mensajes personalizados por cada lead a un archivo de texto."""
    with open(output, "w", encoding="utf-8") as f:
        for i, lead in enumerate(leads, 1):
            mensaje, link = generar_mensaje(lead)
            f.write(f"{'─'*60}\n")
            f.write(f"[{i}] {lead['nombre']} — {lead['categoria']}\n")
            f.write(f"Tel: {lead['telefono']}\n")
            f.write(f"Mensaje:\n{mensaje}\n")
            if link:
                f.write(f"Link WA: {link}\n")
            f.write("\n")
    print(f"Mensajes exportados a: {output}")


if __name__ == "__main__":
    main()
